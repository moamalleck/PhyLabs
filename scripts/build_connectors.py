from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Connectors"

headers = [
    "Name", "Relationship to Ash", "Role / Title", "Organization",
    "Connected To (Communities / Networks)", "Connector Type",
    "Helpfulness History", "Interest in PhysicianLabs", "Priority",
    "Next Action", "Notes"
]

header_fill = PatternFill("solid", start_color="1F4E79")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

ws.row_dimensions[1].height = 36

connectors = [
    [
        "David Jorjani",
        "FI Mentor (Founder Institute)",
        "Startup Mentor / Investor",
        "Founder Institute",
        "FI mentor network; healthcare startup founders; IM physician community (partner is IM resident at top-5 US hospital)",
        "Mentor / Investor Network",
        "Gave detailed 25-min coaching session Apr 28; coined Business Doula framing; offered to stay in touch",
        "High — confirmed he sees something here; offered ongoing engagement",
        "High",
        "Send progress update; ask for introductions to FI mentors with healthcare practice backgrounds",
        "Partner is IM resident at top-5 US hospital — potential future ICP interview. Referenced Alex Hormozi Gym Launch analogy unprompted."
    ],
    [
        "Chelsea Parkman",
        "FI Mentor (Founder Institute)",
        "Business Coach / Startup Mentor",
        "Founder Institute",
        "Profit improvement coaching network (healthcare); independent clinic owners (PT, dental); FI mentor network",
        "Mentor / Connector / Community",
        "Strong positive reaction to pitch; coined business in a box; proactively offered to research healthcare coaching network contacts and pass them to Ash",
        "Very High — immediate resonance; personal client of independent PT clinic and dental practice",
        "High",
        "Follow up to receive promised healthcare coaching network contacts; ask about Vancouver healthcare conference",
        "LinkedIn: https://www.linkedin.com/in/chelseaparkman/ — Has established healthcare connections. Vancouver healthcare conference ~$650 ticket, drivable from Seattle."
    ],
    [
        "KT McBratney",
        "Investor / Pitch Reviewer",
        "Investor / Advisor",
        "Unknown (FI ecosystem)",
        "Investor network; healthcare startup founders; pitch community",
        "Investor / Advisor",
        "Gave detailed narrative pitch feedback May 11; specific and actionable coaching on story structure",
        "Medium — engaged seriously; gave prescriptive fix (one doctor story spine)",
        "Medium",
        "Send updated pitch using Marcus story once refined; ask if she knows investors focused on healthcare tools for independent practitioners",
        "Feedback focused on narrative clarity not viability. Nicole (co-reviewer same session) gave reinforcing feedback."
    ],
    [
        "Marc Krejci",
        "Mentor Panel Reviewer",
        "Startup Mentor",
        "Founder Institute",
        "FI mentor network; startup investor community",
        "Mentor / Investor Network",
        "Reviewed pitch May 20; gave 2/5 on research thoroughness — critical but constructive",
        "Medium — low score reflects pre-validation stage, not dismissal of concept",
        "Medium",
        "Re-engage after first paid customer is secured; update on experiment result",
        "Critical gap he flagged: no paid commitment. First payment resolves his primary objection entirely."
    ],
    [
        "Carmela Fortin",
        "Mentor Panel Reviewer",
        "Startup Mentor",
        "Founder Institute",
        "FI mentor network; startup community",
        "Mentor / Investor Network",
        "Reviewed pitch May 20; gave 8/15 — more favorable; highlighted strong 3-week traction signal",
        "Medium-High — recognized LOIs and design partner signal as meaningful early traction",
        "Medium",
        "Re-engage after first paid customer; mention ICP refinement and Marcus archetype as evidence of learning velocity",
        "Noted three-tier revenue model as well-structured — receptive to business model sophistication."
    ],
    [
        "Saf Malleck, MD",
        "Brother",
        "Orthopedic Surgeon",
        "Independent clinic, Ontario",
        "Ontario physician community; Matthew Tucci (solo ortho); PT in his building (10-person PT practice); Linda Xin (Rejuvenation Dermatology, Oakville); Dr. Phan/Vu (medical finance webinars); junior hospital colleagues; Facebook doctor groups",
        "Physician Network / Family",
        "Highest helpfulness in portfolio — 2 in-depth interviews, WTP signal ($1K/$10K), multiple warm referrals, intro to Dr. Phan/Vu pending; consistently responsive",
        "Very High — confirmed WTP, emotionally invested in the problem, brother of founder",
        "Very High",
        "Ask for warm intro to Matthew Tucci today; confirm Dr. Phan/Vu email received; ask for intro to PT in his building; ask for intro to Linda Xin at Rejuvenation Dermatology",
        "Best source of warm referrals in current network. Has named 5+ referrals not yet contacted. Facebook doctor groups access not yet obtained."
    ],
    [
        "Vince",
        "Mutual friend",
        "Unknown",
        "Unknown",
        "Physician network (connected Ash to Shaun Cowan); likely broader professional network",
        "Personal Network / Connector",
        "Made one warm introduction (Shaun Cowan) proactively",
        "Unknown — not yet engaged on PhysicianLabs specifically",
        "Medium",
        "Re-engage; share what PhysicianLabs is building; ask if he knows other physicians considering independence",
        "Unknown background but demonstrated willingness to make warm physician introductions. Explore his network depth."
    ],
    [
        "Dr. Phan / Dr. Vu",
        "Referred by Saf Malleck",
        "Physician / Medical Finance Educator",
        "Independent (Toronto)",
        "Toronto physician community; medical finance WhatsApp group; physician finance webinar attendees",
        "Physician Community Leader / Educator",
        "Not yet contacted — intro pending from Saf (email not confirmed received)",
        "Unknown — not yet engaged",
        "Very High",
        "Confirm with Saf that email was sent; follow up directly via WhatsApp if contact info obtained",
        "Runs medical finance webinars for physicians. If warm, this is a direct pipeline into the ICP. Highest-priority untapped connector in the network."
    ],
    [
        "Jim Dahle (u/WCInvestor)",
        "Reddit community leader",
        "Physician / Founder / Media",
        "White Coat Investor (whitecoatinvestor.com)",
        "500K+ physician readers; r/whitecoatinvestor (500K+ members); physician financial independence community nationally",
        "Physician Community Leader / Media",
        "No direct contact yet — identified from r/whitecoatinvestor S-corp thread where he commented with resource links",
        "Unknown — not yet contacted",
        "High",
        "Approach via LinkedIn or WCI contact form; offer guest post or resource for physicians asking setup questions; value-first framing only",
        "Highest-reach connector in physician finance/independence space. A WCI newsletter or podcast mention would reach more qualified ICP prospects than any other single channel."
    ],
    [
        "Samir Master, MD MBA",
        "Customer interview subject",
        "Dermatologist / Founder",
        "Dermatology Arts, Bellevue WA",
        "100+ derm practices on ModMed nationally; Pro Alliance (IPA); WWMG (Western Washington Medical Group); Bellevue physician community",
        "Physician Community / Competitor-Adjacent",
        "Gave in-depth interview May 28; introduced Pro Alliance and WWMG as distribution channels; invited Ash to return when there is something concrete to show",
        "Medium — building competing product but invited ongoing relationship; not hostile",
        "High",
        "Return when first prototype or paying customer exists; ask for Pro Alliance and WWMG admin introductions independently",
        "ALERT: Building competing product (Rippling meets Compass for outpatient medicine). Treat as market intelligence source and connector, not prospect. Pro Alliance and WWMG intros are the primary value from this relationship."
    ],
    [
        "Matthew Tucci",
        "Referred by Saf Malleck",
        "Orthopedic Surgeon",
        "Solo practice (then group), Ontario",
        "Ontario physician community; Oshawa physician network; solo-to-group practice owner network",
        "Physician Network",
        "Not yet contacted",
        "Unknown — not yet engaged",
        "High",
        "Ask Saf for warm intro today; frame as: want to learn how Matthew navigated the first year solo before building his group",
        "First-hand experience of the exact transition journey PhysicianLabs is built for. Broke from pyramid scheme clinic and built his own group. Story mirrors Marcus archetype closely."
    ],
    [
        "Pro Alliance Admin",
        "Referred by Samir Master",
        "Association Administrator",
        "Pro Alliance (IPA)",
        "Independent physicians shedding pods from large health systems; Pacific Northwest physician independence community",
        "Physician Association / Distribution Channel",
        "Not yet contacted — name/contact not obtained",
        "Unknown",
        "High",
        "Ask Samir Master for specific admin contact name and email; approach as resource for physicians transitioning out of the IPA",
        "IPAs shedding pods = physicians in active transition = direct pipeline to Marcus archetype. Warm channel if admin intro is secured."
    ],
]

alt_fill = PatternFill("solid", start_color="EBF3FB")
normal_fill = PatternFill("solid", start_color="FFFFFF")
priority_colors = {
    "Very High": "C00000",
    "High": "E26B0A",
    "Medium": "375623",
    "Low": "595959"
}

for row_idx, connector in enumerate(connectors, 2):
    row_fill = alt_fill if row_idx % 2 == 0 else normal_fill
    for col_idx, value in enumerate(connector, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = border
        if col_idx == 9 and value in priority_colors:
            cell.font = Font(name="Arial", size=10, bold=True, color=priority_colors[value])
    ws.row_dimensions[row_idx].height = 80

col_widths = [22, 22, 22, 22, 42, 30, 40, 32, 12, 42, 52]
for col_idx, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = "A2"

notes = wb.create_sheet("Legend")
notes["A1"] = "PhysicianLabs — Connector Tracker"
notes["A1"].font = Font(name="Arial", bold=True, size=14)
notes["A3"] = "Priority Key"
notes["A3"].font = Font(name="Arial", bold=True, size=11)

legend = [
    ("Very High", "Active pipeline or highest-value untapped connector; action this week"),
    ("High", "Strong connector potential; action within 2 weeks"),
    ("Medium", "Useful connector; engage after higher-priority actions are done"),
    ("Low", "Monitor; re-engage when relevant"),
]
for i, (p, desc) in enumerate(legend, 5):
    c = notes.cell(row=i, column=1, value=p)
    c.font = Font(name="Arial", bold=True, color=priority_colors.get(p, "000000"), size=10)
    notes.cell(row=i, column=2, value=desc).font = Font(name="Arial", size=10)

notes["A10"] = "Connector Types"
notes["A10"].font = Font(name="Arial", bold=True, size=11)
types = [
    ("Physician Network", "Physicians with warm referral access to other physicians in or near the ICP"),
    ("Mentor / Investor Network", "FI mentors, advisors, investors who can provide introductions to both customers and capital"),
    ("Physician Community Leader / Educator", "Physicians who run communities, newsletters, podcasts, or webinars for other physicians"),
    ("Physician Association / Distribution Channel", "Associations, IPAs, or physician groups that represent clusters of ICP-fit physicians"),
    ("Personal Network / Connector", "Non-physician contacts with demonstrated willingness to make warm introductions"),
    ("Competitor-Adjacent", "Contacts building in adjacent space; valuable for market intelligence and warm referrals"),
]
for i, (t, desc) in enumerate(types, 12):
    notes.cell(row=i, column=1, value=t).font = Font(name="Arial", bold=True, size=10)
    notes.cell(row=i, column=2, value=desc).font = Font(name="Arial", size=10)

notes.column_dimensions["A"].width = 35
notes.column_dimensions["B"].width = 70

out = r"C:\Users\Admin\Documents\MyNewStartup\docs\business-development\connectors.xlsx"
wb.save(out)
print(f"Saved: {out}")
